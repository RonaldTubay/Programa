"""
App web de carga de reportes de sílabos y programas analíticos.

Uso:
    pip install -r requirements.txt
    python app.py
    Abrir en el navegador: http://localhost:5000
"""

import os
import json
import csv
import io
import tempfile
import traceback
import re
import zipfile
import unicodedata
from functools import wraps
from xml.sax.saxutils import escape

import pandas as pd
import pyodbc
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from flask import Flask, request, render_template, session, redirect, url_for, Response
from werkzeug.security import check_password_hash

# ----------------------------------------------------------------------
# CONFIGURACIÓN
# ----------------------------------------------------------------------
MODO_LOCAL_SIN_LOGIN = True

def _driver_sql_server():
    drivers = pyodbc.drivers()
    if "ODBC Driver 18 for SQL Server" in drivers:
        return "ODBC Driver 18 for SQL Server"
    if "ODBC Driver 17 for SQL Server" in drivers:
        return "ODBC Driver 17 for SQL Server"
    return "ODBC Driver 17 for SQL Server"

CONN_STR = (
    f"DRIVER={{{_driver_sql_server()}}};"
    f"SERVER={os.environ.get('GESTION_SILABOS_SQL_SERVER', '.\\SQLEXPRESS')};"
    "DATABASE=Programa_Silabo;"
    "Trusted_Connection=yes;"
    "TrustServerCertificate=yes;"
)

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 20 * 1024 * 1024

app.secret_key = os.environ.get("FLASK_SECRET_KEY", "CAMBIA-ESTA-CLAVE-POR-UNA-PROPIA-Y-LARGA")

COLUMNAS_REQUERIDAS = [
    "Ref. Facultad", "Facultad", "Ref. Carrera", "Carrera",
    "Ref. Materia", "Cod. Materia", "Materia", "Paralelo",
    "Profesor del Curso", "Profesor Sílabo", "Existe Sílabo", "Finalizado",
]

COLUMNAS_REQUERIDAS_ANALITICO = [
    "Ref. Facultad", "Facultad", "Ref. Carrera", "Carrera", "Nivel",
    "Ref. Materia", "Cod. Materia", "Materia", "Régimen Materia",
    "Ref. Ultimo Plan Analítico", "Existe Plan Analítico", "Observación",
]

def iniciales_usuario():
    nombre = session.get("nombre_completo", "Usuario Local")
    return "".join(p[0] for p in nombre.split()[:2]).upper()

def login_requerido(vista):
    @wraps(vista)
    def envoltura(*args, **kwargs):
        if MODO_LOCAL_SIN_LOGIN:
            session.setdefault("nombre_usuario", "local")
            session.setdefault("nombre_completo", "Usuario Local")
            return vista(*args, **kwargs)
        if "nombre_usuario" not in session:
            return redirect(url_for("login", siguiente=request.path))
        return vista(*args, **kwargs)
    return envoltura

def _obtener_alertas_resumen():
    if not MODO_LOCAL_SIN_LOGIN and "nombre_usuario" not in session:
        return {"total": 0}
    try:
        conn = pyodbc.connect(CONN_STR)
        cur = conn.cursor()
        cur.execute("""
            SELECT
                SUM(CASE WHEN s.existe_silabo = 0 OR s.existe_silabo IS NULL THEN 1 ELSE 0 END),
                SUM(CASE WHEN s.existe_silabo = 1 AND (s.finalizado = 0 OR s.finalizado IS NULL) THEN 1 ELSE 0 END)
            FROM OfertaAcademica oa
            LEFT JOIN Silabo s ON s.oferta_id = oa.oferta_id
        """)
        sin_silabo, por_finalizar = cur.fetchone()
        cur.execute("""
            SELECT COUNT(*)
            FROM Materia m
            LEFT JOIN PlanAnalitico pa ON pa.materia_id = m.materia_id
            WHERE pa.existe_plan_analitico = 0 OR pa.existe_plan_analitico IS NULL
        """)
        sin_plan = cur.fetchone()[0] or 0
        cur.execute("""
            SELECT COUNT(*)
            FROM CargaArchivo
            WHERE estado = 'ERROR'
              AND fecha_carga >= DATEADD(day, -30, SYSDATETIME())
        """)
        cargas_error = cur.fetchone()[0] or 0
        conn.close()
        sin_silabo = sin_silabo or 0
        por_finalizar = por_finalizar or 0
        total = sin_silabo + por_finalizar + sin_plan + cargas_error
        firma = json.dumps(
            {
                "sin_silabo": sin_silabo,
                "por_finalizar": por_finalizar,
                "sin_plan": sin_plan,
                "cargas_error": cargas_error,
            },
            sort_keys=True,
        )
        total_visible = 0 if session.get("alertas_revisadas_firma") == firma else total
        return {
            "total": total_visible,
            "total_operativo": total,
            "sin_silabo": sin_silabo,
            "por_finalizar": por_finalizar,
            "sin_plan": sin_plan,
            "cargas_error": cargas_error,
            "firma": firma,
        }
    except Exception:
        return {"total": 0}

@app.context_processor
def inyectar_alertas():
    return {"alertas_resumen": _obtener_alertas_resumen()}

def clasificar_tipo_unidad(nombre_archivo, unidad_academica=None):
    texto = f"{nombre_archivo} {unidad_academica or ''}".lower()
    sedes_extensiones = [
        "bahía", "bahia", "chone", "el carmen", "pedernales",
        "tosagua", "sucre", "jama", "jipijapa", "pichincha",
    ]
    if "campus" in texto: return "CAMPUS"
    if "extension" in texto or "extensión" in texto: return "EXTENSION"
    if "sede" in texto: return "SEDE"
    if any(nombre in texto for nombre in sedes_extensiones): return "EXTENSION"
    return "FACULTAD"

def limpio(v):
    if v is None:
        return None
    try:
        if pd.isna(v):
            return None
    except (TypeError, ValueError):
        return None
    v = str(v).strip()
    return v if v else None

def limpiar_referencia(v):
    v = limpio(v)
    if v is None:
        return None
    try:
        numero = float(v)
        if numero.is_integer():
            return str(int(numero))
    except ValueError:
        pass
    return v

def clave_texto(valor):
    valor = limpio(valor) or ""
    valor = unicodedata.normalize("NFKD", valor)
    valor = "".join(c for c in valor if not unicodedata.combining(c))
    valor = re.sub(r"[^A-Za-z0-9]+", " ", valor).upper()
    return re.sub(r"\s+", " ", valor).strip()

def titulo_academico(valor):
    valor = re.sub(r"\s+", " ", str(valor or "")).strip()
    menores = {"DE", "DEL", "LA", "LAS", "LOS", "Y", "EN", "A"}
    partes = []
    for palabra in valor.upper().split():
        partes.append(palabra.lower() if palabra in menores else palabra.capitalize())
    return " ".join(partes)

UNIDADES_EQUIVALENTES = {
    "JUNIN": "Campus Junin",
    "CAMPUS JUNIN": "Campus Junin",
    "BAHIA": "Extension Bahia de Caraquez",
    "BAHIA DE CARAQUEZ": "Extension Bahia de Caraquez",
    "EXTENSION BAHIA": "Extension Bahia de Caraquez",
    "EXTENSION BAHIA DE CARAQUEZ": "Extension Bahia de Caraquez",
    "CHONE": "Extension Chone",
    "EXTENSION CHONE": "Extension Chone",
    "EL CARMEN": "Extension El Carmen",
    "EXTENSION EL CARMEN": "Extension El Carmen",
    "PEDERNALES": "Extension Pedernales",
    "EXTENSION PEDERNALES": "Extension Pedernales",
    "PICHINCHA": "Campus Pichincha",
    "CAMPUS PICHINCHA": "Campus Pichincha",
}

def normalizar_nombre_unidad(nombre, tipo=None):
    clave = clave_texto(nombre)
    if clave in UNIDADES_EQUIVALENTES:
        return UNIDADES_EQUIVALENTES[clave]
    if tipo == "CAMPUS" and not clave.startswith("CAMPUS "):
        return f"Campus {titulo_academico(nombre)}"
    if tipo == "EXTENSION" and not clave.startswith("EXTENSION "):
        return f"Extension {titulo_academico(nombre)}"
    if tipo == "SEDE" and not clave.startswith("SEDE "):
        return f"Sede {titulo_academico(nombre)}"
    return titulo_academico(nombre)

def normalizar_nombre_carrera(nombre_original):
    nombre = re.sub(r"\s+", " ", str(nombre_original or "")).strip()
    nombre = re.sub(r"^(CAMPUS|EXTENSION|EXTENSIÓN|SEDE)\s+", "", nombre, flags=re.IGNORECASE).strip()
    nombre = re.sub(r"\s*[-–]\s*(CAMPUS|EXTENSION|EXTENSIÓN|SEDE)\s+\w+\s*$", "", nombre, flags=re.IGNORECASE).strip()
    return nombre

def extraer_malla_carrera(nombre_original):
    nombre = normalizar_nombre_carrera(nombre_original)
    malla = None
    match = re.search(r'(\b20\d{2}(?:\s*[-A-ZÑÁÉÍÓÚ0-9()]+.*)?|\(Créditos\).*)$', nombre, flags=re.IGNORECASE)
    if match:
        malla = match.group(1).strip()
        nombre = nombre[:match.start()].strip()
        nombre = re.sub(r'[-_\s]+$', '', nombre).strip()
    return nombre, malla

# ----------------------------------------------------------------------
# LECTURA Y PARSEO DE REPORTES
# ----------------------------------------------------------------------
def detectar_tipo_reporte_por_columnas(datos):
    columnas = {str(c).strip() for c in datos.columns}
    columnas_silabos = set(COLUMNAS_REQUERIDAS)
    columnas_asignaturas = set(COLUMNAS_REQUERIDAS_ANALITICO)

    if columnas_silabos.issubset(columnas):
        return "SILABOS"
    if columnas_asignaturas.issubset(columnas):
        return "ASIGNATURAS"
    if columnas & {"Profesor Sílabo", "Existe Sílabo", "Finalizado"}:
        return "SILABOS_INCOMPLETO"
    if columnas & {"Nivel", "Régimen Materia", "Existe Plan Analítico", "Ref. Ultimo Plan Analítico"}:
        return "ASIGNATURAS_INCOMPLETO"
    return "DESCONOCIDO"

def diagnosticar_archivo_carga(filepath, tipo_esperado):
    esperadas = COLUMNAS_REQUERIDAS if tipo_esperado == "SILABOS" else COLUMNAS_REQUERIDAS_ANALITICO
    etiqueta_esperada = "Silabos" if tipo_esperado == "SILABOS" else "Asignaturas"
    try:
        tablas = pd.read_html(filepath)
    except Exception as exc:
        return {
            "valido": False,
            "mensaje": "No se pudo leer el archivo como reporte Excel exportado.",
            "detalle": str(exc),
            "esperado": etiqueta_esperada,
        }

    if len(tablas) < 2:
        return {
            "valido": False,
            "mensaje": "El archivo no tiene la estructura esperada.",
            "detalle": "Se esperaba una tabla de encabezado y una tabla de datos.",
            "esperado": etiqueta_esperada,
        }

    datos = tablas[1].copy()
    datos.columns = [str(c).strip() for c in datos.columns]
    columnas = list(datos.columns)
    tipo_detectado = detectar_tipo_reporte_por_columnas(datos)
    faltantes = [c for c in esperadas if c not in columnas]
    claves = ["Ref. Carrera", "Ref. Materia"]
    if tipo_esperado == "SILABOS":
        claves.append("Paralelo")
    columnas_clave = [c for c in claves if c in columnas]
    filas_sin_materia = 0
    duplicados = 0
    if "Ref. Materia" in columnas:
        filas_sin_materia = int(datos["Ref. Materia"].isna().sum())
    if columnas_clave:
        duplicados = int(datos.duplicated(subset=columnas_clave, keep=False).sum())

    tipo_cruzado = (
        tipo_esperado == "SILABOS" and tipo_detectado == "ASIGNATURAS"
    ) or (
        tipo_esperado == "ASIGNATURAS" and tipo_detectado == "SILABOS"
    )
    valido = not faltantes and not tipo_cruzado
    if tipo_cruzado:
        mensaje = "El archivo pertenece a la otra seccion del sistema."
    elif faltantes:
        mensaje = "Faltan columnas obligatorias para procesar el archivo."
    else:
        mensaje = "El archivo cumple con las columnas requeridas."

    return {
        "valido": valido,
        "mensaje": mensaje,
        "esperado": etiqueta_esperada,
        "detectado": tipo_detectado.replace("_", " ").title(),
        "filas": int(len(datos)),
        "columnas": len(columnas),
        "faltantes": faltantes,
        "filas_sin_materia": filas_sin_materia,
        "duplicados": duplicados,
    }

def parsear_reporte(filepath):
    tablas = pd.read_html(filepath)
    if len(tablas) < 2:
        raise ValueError("El archivo no tiene el formato esperado.")

    meta, datos = tablas[0], tablas[1]
    periodo = unidad_academica = None
    for val in meta.values.flatten():
        val = str(val)
        if "Período:" in val and periodo is None:
            periodo = val.split(":", 1)[1].strip()
        elif "Unidad Académica:" in val and unidad_academica is None:
            unidad_academica = val.split(":", 1)[1].strip()

    if not periodo or not unidad_academica:
        raise ValueError("Falta el Período o la Unidad Académica.")

    datos.columns = [str(c).strip() for c in datos.columns]
    tipo_detectado = detectar_tipo_reporte_por_columnas(datos)
    if tipo_detectado == "ASIGNATURAS":
        raise ValueError("Este archivo corresponde a Asignaturas. Cárgalo desde la sección Asignaturas.")

    faltantes = [c for c in COLUMNAS_REQUERIDAS if c not in datos.columns]
    if faltantes:
        raise ValueError(f"El archivo no corresponde al formato de Sílabos. Faltan columnas: {', '.join(faltantes)}")

    if datos.empty:
        return periodo, unidad_academica, datos

    for col in ["Ref. Facultad", "Ref. Carrera", "Ref. Materia"]:
        datos[col] = datos[col].astype("Int64").astype(str)

    datos["Existe Sílabo"] = datos["Existe Sílabo"].astype(str).str.strip().str.lower().eq("sí")
    datos["Finalizado"] = datos["Finalizado"].apply(lambda v: bool(pd.notna(v) and float(v) == 1.0))
    return periodo, unidad_academica, datos


def parsear_reporte_analitico(filepath):
    tablas = pd.read_html(filepath)
    if len(tablas) < 2:
        raise ValueError("El archivo no tiene el formato esperado.")

    meta, datos = tablas[0], tablas[1]
    periodo = unidad_academica = None
    for val in meta.values.flatten():
        val = str(val)
        if "Período:" in val and periodo is None:
            periodo = val.split(":", 1)[1].strip()
        elif "Unidad Académica:" in val and unidad_academica is None:
            unidad_academica = val.split(":", 1)[1].strip()

    if not periodo or not unidad_academica:
        raise ValueError("No se pudo leer el Período o la Unidad Académica.")

    datos.columns = [str(c).strip() for c in datos.columns]
    tipo_detectado = detectar_tipo_reporte_por_columnas(datos)
    if tipo_detectado == "SILABOS":
        raise ValueError("Este archivo corresponde a Sílabos. Cárgalo desde la sección Sílabos.")

    faltantes = [c for c in COLUMNAS_REQUERIDAS_ANALITICO if c not in datos.columns]
    if faltantes:
        raise ValueError(f"El archivo no corresponde al formato de Asignaturas. Faltan columnas: {', '.join(faltantes)}")

    if datos.empty: return periodo, unidad_academica, datos

    for col in ["Ref. Facultad", "Ref. Carrera", "Ref. Materia"]:
        datos[col] = datos[col].astype("Int64").astype(str)

    datos["Nivel"] = datos["Nivel"].astype("Int64")
    datos["Existe Plan Analítico"] = datos["Existe Plan Analítico"].astype(str).str.strip().str.lower().eq("si")
    datos["Ref. Ultimo Plan Analítico"] = datos["Ref. Ultimo Plan Analítico"].apply(limpiar_referencia)
    if "Período Académico" in datos.columns:
        datos["_ofertada_este_periodo"] = datos["Período Académico"].notna()
    else:
        datos["_ofertada_este_periodo"] = False

    return periodo, unidad_academica, datos

# ----------------------------------------------------------------------
# HELPERS DE BASE DE DATOS
# ----------------------------------------------------------------------
def get_or_create_unidad(cur, ref_facultad, nombre, tipo):
    nombre_normalizado = normalizar_nombre_unidad(nombre, tipo)
    cur.execute("SELECT unidad_academica_id, tipo, nombre FROM UnidadAcademica WHERE ref_facultad = ?", ref_facultad)
    row = cur.fetchone()
    if row:
        unidad_id, tipo_actual, nombre_actual = row
        if tipo_actual == "FACULTAD" and tipo in ("SEDE", "EXTENSION", "CAMPUS"):
            cur.execute("UPDATE UnidadAcademica SET tipo = ? WHERE unidad_academica_id = ?", tipo, unidad_id)
        if nombre_actual != nombre_normalizado:
            cur.execute("UPDATE UnidadAcademica SET nombre = ? WHERE unidad_academica_id = ?", nombre_normalizado, unidad_id)
        return unidad_id
    cur.execute(
        "INSERT INTO UnidadAcademica (ref_facultad, nombre, tipo) OUTPUT INSERTED.unidad_academica_id VALUES (?, ?, ?)",
        ref_facultad, nombre_normalizado, tipo,
    )
    return cur.fetchone()[0]

def get_or_create_carrera(cur, ref_carrera, nombre_original, unidad_academica_id):
    nombre_limpio, malla_extraida = extraer_malla_carrera(nombre_original)
    cur.execute("SELECT carrera_id, nombre, malla FROM Carrera WHERE ref_carrera = ?", ref_carrera)
    row = cur.fetchone()
    if row:
        carrera_id, nombre_actual, malla_actual = row
        if (malla_extraida and not malla_actual) or (nombre_limpio != nombre_actual):
            cur.execute(
                "UPDATE Carrera SET nombre = ?, malla = COALESCE(malla, ?) WHERE carrera_id = ?",
                nombre_limpio, malla_extraida, carrera_id
            )
        return carrera_id

    cur.execute(
        "INSERT INTO Carrera (unidad_academica_id, ref_carrera, nombre, malla) OUTPUT INSERTED.carrera_id VALUES (?, ?, ?, ?)",
        unidad_academica_id, ref_carrera, nombre_limpio, malla_extraida,
    )
    return cur.fetchone()[0]

def get_or_create_materia(cur, ref_materia, cod_materia, nombre, carrera_id, nivel=None, regimen_materia=None):
    cur.execute(
        "SELECT materia_id, nivel, regimen_materia FROM Materia WHERE carrera_id = ? AND ref_materia = ?",
        carrera_id, ref_materia,
    )
    row = cur.fetchone()
    if row:
        materia_id, nivel_actual, regimen_actual = row
        if (nivel is not None and nivel_actual is None) or (regimen_materia is not None and regimen_actual is None):
            cur.execute(
                "UPDATE Materia SET nivel = COALESCE(nivel, ?), regimen_materia = COALESCE(regimen_materia, ?) WHERE materia_id = ?",
                nivel, regimen_materia, materia_id,
            )
        return materia_id
    cur.execute(
        "INSERT INTO Materia (carrera_id, ref_materia, cod_materia, nombre, nivel, regimen_materia) OUTPUT INSERTED.materia_id VALUES (?, ?, ?, ?, ?, ?)",
        carrera_id, ref_materia, cod_materia, nombre, nivel, regimen_materia,
    )
    return cur.fetchone()[0]

def get_or_create_periodo(cur, nombre):
    cur.execute("SELECT periodo_id FROM PeriodoAcademico WHERE nombre = ?", nombre)
    row = cur.fetchone()
    if row: return row[0]
    cur.execute("INSERT INTO PeriodoAcademico (nombre) OUTPUT INSERTED.periodo_id VALUES (?)", nombre)
    return cur.fetchone()[0]

def get_or_create_docente(cur, nombre_completo):
    if nombre_completo is None: return None
    cur.execute("SELECT docente_id FROM Docente WHERE nombre_completo = ?", nombre_completo)
    row = cur.fetchone()
    if row: return row[0]
    cur.execute("INSERT INTO Docente (nombre_completo) OUTPUT INSERTED.docente_id VALUES (?)", nombre_completo)
    return cur.fetchone()[0]

def get_or_create_oferta(cur, materia_id, periodo_id, paralelo, profesor_curso_id, profesor_silabo_id):
    if profesor_silabo_id is None:
        cur.execute(
            "SELECT oferta_id FROM OfertaAcademica WHERE materia_id = ? AND periodo_id = ? AND paralelo = ? AND profesor_silabo_id IS NULL",
            materia_id, periodo_id, paralelo,
        )
    else:
        cur.execute(
            "SELECT oferta_id FROM OfertaAcademica WHERE materia_id = ? AND periodo_id = ? AND paralelo = ? AND profesor_silabo_id = ?",
            materia_id, periodo_id, paralelo, profesor_silabo_id,
        )
    row = cur.fetchone()
    if row: return row[0]
    cur.execute(
        "INSERT INTO OfertaAcademica (materia_id, periodo_id, paralelo, profesor_curso_id, profesor_silabo_id) OUTPUT INSERTED.oferta_id VALUES (?, ?, ?, ?, ?)",
        materia_id, periodo_id, paralelo, profesor_curso_id, profesor_silabo_id,
    )
    return cur.fetchone()[0]

# ----------------------------------------------------------------------
# SINCRONIZACIÓN Y ESTADOS
# ----------------------------------------------------------------------
def sincronizar_silabo(cur, oferta_id, existe_silabo_nuevo, finalizado_nuevo, usuario):
    cur.execute("SELECT silabo_id, existe_silabo, finalizado, version_actual FROM Silabo WHERE oferta_id = ?", oferta_id)
    row = cur.fetchone()

    if row is None:
        cur.execute("INSERT INTO Silabo (oferta_id, existe_silabo, finalizado) VALUES (?, ?, ?)", oferta_id, existe_silabo_nuevo, finalizado_nuevo)
        return "creado"

    silabo_id, existe_anterior, finalizado_anterior, version_actual = row
    if bool(existe_anterior) == existe_silabo_nuevo and bool(finalizado_anterior) == finalizado_nuevo:
        return "sin_cambios"

    cur.execute(
        "INSERT INTO SilaboHistorial (silabo_id, version, existe_silabo_anterior, existe_silabo_nuevo, finalizado_anterior, finalizado_nuevo, usuario_responsable) VALUES (?, ?, ?, ?, ?, ?, ?)",
        silabo_id, version_actual, existe_anterior, existe_silabo_nuevo, finalizado_anterior, finalizado_nuevo, usuario,
    )
    cur.execute(
        "UPDATE Silabo SET existe_silabo = ?, finalizado = ?, version_actual = version_actual + 1, fecha_registro = SYSDATETIME() WHERE silabo_id = ?",
        existe_silabo_nuevo, finalizado_nuevo, silabo_id,
    )
    return "actualizado"

def sincronizar_plan_analitico(cur, materia_id, existe_nuevo, ref_plan_nuevo, observacion_nueva, carga_id):
    ref_plan_nuevo = limpiar_referencia(ref_plan_nuevo)
    observacion_nueva = limpio(observacion_nueva)

    cur.execute("SELECT plan_analitico_id, existe_plan_analitico, ref_ultimo_plan_analitico, version_actual FROM PlanAnalitico WHERE materia_id = ?", materia_id)
    row = cur.fetchone()

    if row is None:
        cur.execute("INSERT INTO PlanAnalitico (materia_id, existe_plan_analitico, ref_ultimo_plan_analitico, observacion) VALUES (?, ?, ?, ?)", materia_id, existe_nuevo, ref_plan_nuevo, observacion_nueva)
        return "creado"

    plan_id, existe_anterior, ref_anterior, version_actual = row
    if bool(existe_anterior) == existe_nuevo and ref_anterior == ref_plan_nuevo:
        return "sin_cambios"

    cur.execute(
        "INSERT INTO PlanAnaliticoHistorial (plan_analitico_id, version, existe_anterior, existe_nuevo, ref_plan_anterior, ref_plan_nuevo, carga_id) VALUES (?, ?, ?, ?, ?, ?, ?)",
        plan_id, version_actual, existe_anterior, existe_nuevo, ref_anterior, ref_plan_nuevo, carga_id,
    )
    cur.execute(
        "UPDATE PlanAnalitico SET existe_plan_analitico = ?, ref_ultimo_plan_analitico = ?, observacion = ?, version_actual = version_actual + 1, fecha_registro = SYSDATETIME() WHERE plan_analitico_id = ?",
        existe_nuevo, ref_plan_nuevo, observacion_nueva, plan_id,
    )
    return "actualizado"

def registrar_estado_malla(cur, materia_id, carga_id, periodo_texto, ofertada):
    cur.execute(
        "INSERT INTO MateriaMallaEstado (materia_id, carga_id, periodo_academico_texto, ofertada_este_periodo) VALUES (?, ?, ?, ?)",
        materia_id, carga_id, periodo_texto, ofertada,
    )

def registrar_carga(cur, nombre_archivo, periodo, unidad, contadores, estado, mensaje_error=None):
    cur.execute(
        "INSERT INTO CargaArchivo (nombre_archivo, periodo, unidad_academica, registros_creados, registros_actualizados, registros_sin_cambio, estado, mensaje_error, tipo_reporte) VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'SILABO')",
        nombre_archivo, periodo, unidad, contadores.get("creado") if contadores else None,
        contadores.get("actualizado") if contadores else None, contadores.get("sin_cambios") if contadores else None,
        estado, mensaje_error,
    )

def crear_carga_pendiente(cur, nombre_archivo, tipo_reporte):
    cur.execute(
        "INSERT INTO CargaArchivo (nombre_archivo, estado, tipo_reporte) OUTPUT INSERTED.carga_id VALUES (?, 'EN_PROCESO', ?)",
        nombre_archivo, tipo_reporte,
    )
    return cur.fetchone()[0]

def finalizar_carga(cur, carga_id, periodo, unidad, contadores, estado, mensaje_error=None):
    cur.execute(
        "UPDATE CargaArchivo SET periodo = ?, unidad_academica = ?, registros_creados = ?, registros_actualizados = ?, registros_sin_cambio = ?, estado = ?, mensaje_error = ? WHERE carga_id = ?",
        periodo, unidad, contadores.get("creado") if contadores else None, contadores.get("actualizado") if contadores else None,
        contadores.get("sin_cambios") if contadores else None, estado, mensaje_error, carga_id,
    )

def obtener_cargas_recientes(cur, limite=10, tipo_reporte=None):
    if tipo_reporte:
        cur.execute(f"""
            SELECT TOP {limite} nombre_archivo, fecha_carga, periodo, unidad_academica,
                   registros_creados, registros_actualizados, registros_sin_cambio, estado
            FROM CargaArchivo WHERE tipo_reporte = ? ORDER BY fecha_carga DESC
        """, tipo_reporte)
    else:
        cur.execute(f"""
            SELECT TOP {limite} nombre_archivo, fecha_carga, periodo, unidad_academica,
                   registros_creados, registros_actualizados, registros_sin_cambio, estado
            FROM CargaArchivo ORDER BY fecha_carga DESC
        """)
    columnas = [c[0] for c in cur.description]
    return [dict(zip(columnas, fila)) for fila in cur.fetchall()]

def registrar_error_validacion(nombre_archivo, tipo_reporte, mensaje):
    try:
        conn = pyodbc.connect(CONN_STR)
        cur = conn.cursor()
        if tipo_reporte == "SILABO":
            registrar_carga(cur, nombre_archivo, None, None, None, "ERROR", mensaje)
        else:
            carga_id = crear_carga_pendiente(cur, nombre_archivo, tipo_reporte)
            finalizar_carga(cur, carga_id, None, None, None, "ERROR", mensaje)
        conn.commit()
        conn.close()
    except Exception:
        traceback.print_exc()

# ----------------------------------------------------------------------
# FLUJOS DE MIGRACIÓN
# ----------------------------------------------------------------------
def migrar_archivo(cur, filepath, nombre_archivo_original, usuario):
    periodo, unidad_academica, datos = parsear_reporte(filepath)
    contadores = {"creado": 0, "actualizado": 0, "sin_cambios": 0}

    if datos.empty: return periodo, unidad_academica, contadores

    tipo_unidad = clasificar_tipo_unidad(nombre_archivo_original, unidad_academica)
    periodo_id = get_or_create_periodo(cur, periodo)

    for _, fila in datos.iterrows():
        unidad_id = get_or_create_unidad(cur, fila["Ref. Facultad"], fila["Facultad"].strip(), tipo_unidad)
        carrera_id = get_or_create_carrera(cur, fila["Ref. Carrera"], fila["Carrera"].strip(), unidad_id)
        materia_id = get_or_create_materia(cur, fila["Ref. Materia"], fila["Cod. Materia"].strip(), fila["Materia"].strip(), carrera_id)

        docente_curso_id = get_or_create_docente(cur, limpio(fila["Profesor del Curso"]))
        docente_silabo_id = get_or_create_docente(cur, limpio(fila["Profesor Sílabo"]))

        oferta_id = get_or_create_oferta(cur, materia_id, periodo_id, fila["Paralelo"].strip(), docente_curso_id, docente_silabo_id)

        resultado = sincronizar_silabo(cur, oferta_id, bool(fila["Existe Sílabo"]), bool(fila["Finalizado"]), usuario)
        contadores[resultado] += 1

    return periodo, unidad_academica, contadores


def migrar_archivo_analitico(cur, filepath, nombre_archivo_original, carga_id):
    periodo, unidad_academica, datos = parsear_reporte_analitico(filepath)
    contadores = {"creado": 0, "actualizado": 0, "sin_cambios": 0}

    if datos.empty:
        return periodo, unidad_academica, contadores

    tipo_unidad = clasificar_tipo_unidad(nombre_archivo_original, unidad_academica)
    tiene_periodo_col = "Período Académico" in datos.columns

    for _, fila in datos.iterrows():
        unidad_id = get_or_create_unidad(cur, fila["Ref. Facultad"], fila["Facultad"].strip(), tipo_unidad)
        carrera_id = get_or_create_carrera(cur, fila["Ref. Carrera"], fila["Carrera"].strip(), unidad_id)
        materia_id = get_or_create_materia(
            cur, fila["Ref. Materia"], fila["Cod. Materia"].strip(), fila["Materia"].strip(),
            carrera_id,
            nivel=int(fila["Nivel"]) if pd.notna(fila["Nivel"]) else None,
            regimen_materia=limpio(fila["Régimen Materia"]),
        )

        resultado = sincronizar_plan_analitico(
            cur, materia_id, bool(fila["Existe Plan Analítico"]), fila["Ref. Ultimo Plan Analítico"],
            limpio(fila["Observación"]), carga_id,
        )
        contadores[resultado] += 1

        registrar_estado_malla(
            cur, materia_id, carga_id, limpio(fila["Período Académico"]) if tiene_periodo_col else None,
            bool(fila["_ofertada_este_periodo"]),
        )

    return periodo, unidad_academica, contadores

# ----------------------------------------------------------------------
# RUTAS DE AUTENTICACIÓN
# ----------------------------------------------------------------------
@app.route("/login", methods=["GET", "POST"])
def login():
    if MODO_LOCAL_SIN_LOGIN:
        session.setdefault("nombre_usuario", "local")
        session.setdefault("nombre_completo", "Usuario Local")
        return redirect(url_for("dashboard"))
    error = None
    if request.method == "POST":
        nombre_usuario = request.form.get("nombre_usuario", "").strip()
        password = request.form.get("password", "")

        conn = pyodbc.connect(CONN_STR)
        cur = conn.cursor()
        cur.execute("SELECT nombre_usuario, hash_password, nombre_completo FROM Usuario WHERE nombre_usuario = ? AND activo = 1", nombre_usuario)
        row = cur.fetchone()
        conn.close()

        if row and check_password_hash(row.hash_password, password):
            session["nombre_usuario"] = row.nombre_usuario
            session["nombre_completo"] = row.nombre_completo
            siguiente = request.args.get("siguiente") or url_for("index")
            return redirect(siguiente)
        error = "Usuario o contraseña incorrectos."

    return render_template("login.html", error=error)

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("dashboard" if MODO_LOCAL_SIN_LOGIN else "login"))

@app.route("/alertas/marcar-revisadas", methods=["POST"])
@login_requerido
def marcar_alertas_revisadas():
    resumen = _obtener_alertas_resumen()
    if resumen.get("firma"):
        session["alertas_revisadas_firma"] = resumen["firma"]
    return redirect(url_for("alertas"))

@app.route("/alertas", methods=["GET"])
@login_requerido
def alertas():
    conn = pyodbc.connect(CONN_STR)
    cur = conn.cursor()

    cur.execute("""
        SELECT COUNT(*)
        FROM OfertaAcademica oa
        LEFT JOIN Silabo s ON s.oferta_id = oa.oferta_id
        WHERE s.existe_silabo = 0 OR s.existe_silabo IS NULL
    """)
    sin_silabo = cur.fetchone()[0] or 0

    cur.execute("""
        SELECT COUNT(*)
        FROM OfertaAcademica oa
        LEFT JOIN Silabo s ON s.oferta_id = oa.oferta_id
        WHERE s.existe_silabo = 1 AND (s.finalizado = 0 OR s.finalizado IS NULL)
    """)
    por_finalizar = cur.fetchone()[0] or 0

    cur.execute("""
        SELECT COUNT(*)
        FROM Materia m
        LEFT JOIN PlanAnalitico pa ON pa.materia_id = m.materia_id
        WHERE pa.existe_plan_analitico = 0 OR pa.existe_plan_analitico IS NULL
    """)
    sin_plan = cur.fetchone()[0] or 0

    cur.execute("""
        SELECT COUNT(*)
        FROM CargaArchivo
        WHERE estado = 'ERROR'
          AND fecha_carga >= DATEADD(day, -30, SYSDATETIME())
    """)
    cargas_error = cur.fetchone()[0] or 0

    cur.execute("""
        SELECT TOP 8 nombre_archivo, fecha_carga, periodo, unidad_academica, mensaje_error, tipo_reporte
        FROM CargaArchivo
        WHERE estado = 'ERROR'
        ORDER BY fecha_carga DESC
    """)
    errores = [
        {
            "archivo": r[0],
            "fecha": r[1],
            "periodo": r[2],
            "unidad": r[3],
            "mensaje": r[4],
            "tipo": r[5],
        }
        for r in cur.fetchall()
    ]

    cur.execute("""
        SELECT TOP 2 carga_id, nombre_archivo, fecha_carga, periodo, unidad_academica
        FROM CargaArchivo
        WHERE tipo_reporte = 'PROGRAMA_ANALITICO' AND estado = 'EXITOSO'
        ORDER BY fecha_carga DESC
    """)
    cargas_asignaturas = cur.fetchall()
    cambios_disponibles = len(cargas_asignaturas) >= 2

    conn.close()
    tarjetas = [
        {
            "titulo": "Sílabos pendientes",
            "total": sin_silabo,
            "descripcion": "Ofertas académicas sin sílabo registrado.",
            "url": "/silabos/reporte?estado=sin_silabo",
            "color": "roja",
        },
        {
            "titulo": "Sílabos por finalizar",
            "total": por_finalizar,
            "descripcion": "Sílabos existentes que aún no constan como finalizados.",
            "url": "/silabos/reporte?estado=por_finalizar",
            "color": "amarilla",
        },
        {
            "titulo": "Asignaturas sin plan",
            "total": sin_plan,
            "descripcion": "Asignaturas sin plan analítico registrado.",
            "url": "/asignaturas/reporte?estado=sin_plan",
            "color": "roja",
        },
        {
            "titulo": "Cambios de malla",
            "total": 1 if cambios_disponibles else 0,
            "descripcion": "Hay cargas suficientes para comparar asignaturas entre versiones.",
            "url": "/asignaturas/cambios",
            "color": "verde" if cambios_disponibles else "amarilla",
        },
        {
            "titulo": "Errores de carga",
            "total": cargas_error,
            "descripcion": "Archivos con errores durante los ultimos 30 dias.",
            "url": "/historial",
            "color": "roja" if cargas_error else "verde",
        },
    ]
    return render_template(
        "alertas.html", activo="alertas", iniciales=iniciales_usuario(),
        tarjetas=tarjetas, errores=errores,
    )

# ----------------------------------------------------------------------
# RUTAS DEL DASHBOARD (Incorpora el formateo visual de la malla)
# ----------------------------------------------------------------------
@app.route("/dashboard", methods=["GET"])
@login_requerido
def dashboard():
    filtros = {
        "modo": request.args.get("modo", "general"),
        "periodo": request.args.get("periodo", ""),
        "tipo_unidad": request.args.get("tipo_unidad", ""),
        "unidad_id": request.args.get("unidad_id", ""),
        "carrera_id": request.args.get("carrera_id", ""),
        "estado": request.args.get("estado", ""),
        "buscar": request.args.get("buscar", ""),
    }
    pagina = max(1, request.args.get("pagina", 1, type=int))
    por_pagina = 50

    conn = pyodbc.connect(CONN_STR)
    cur = conn.cursor()

    where_sql, params = _construir_filtro_dashboard(filtros)
    base_from = """
        FROM Materia m
        JOIN Carrera c           ON c.carrera_id = m.carrera_id
        JOIN UnidadAcademica ua  ON ua.unidad_academica_id = c.unidad_academica_id
        LEFT JOIN OfertaAcademica oa ON oa.materia_id = m.materia_id
        LEFT JOIN PeriodoAcademico p ON p.periodo_id = oa.periodo_id
        LEFT JOIN Docente dc     ON dc.docente_id = oa.profesor_curso_id
        LEFT JOIN Docente ds     ON ds.docente_id = oa.profesor_silabo_id
        LEFT JOIN Silabo s       ON s.oferta_id = oa.oferta_id
        LEFT JOIN PlanAnalitico pa ON pa.materia_id = m.materia_id
        OUTER APPLY (
            SELECT TOP 1 estado.periodo_academico_texto, estado.ofertada_este_periodo,
                         ca.periodo AS periodo_carga
            FROM MateriaMallaEstado estado
            JOIN CargaArchivo ca ON ca.carga_id = estado.carga_id
            WHERE estado.materia_id = m.materia_id
            ORDER BY estado.fecha_registro DESC, estado.estado_id DESC
        ) mme
    """

    # --- KPIs ---
    cur.execute(f"""
        SELECT COUNT(oa.oferta_id),
               SUM(CASE WHEN s.existe_silabo = 1 THEN 1 ELSE 0 END),
               SUM(CASE WHEN s.finalizado = 1 THEN 1 ELSE 0 END),
               COUNT(DISTINCT m.materia_id),
               COUNT(DISTINCT CASE WHEN pa.existe_plan_analitico = 1 THEN m.materia_id END),
               COUNT(DISTINCT c.carrera_id)
        {base_from} {where_sql}
    """, params)
    total, con_silabo, finalizados, total_asignaturas, con_plan, total_carreras = cur.fetchone()
    total = total or 0
    con_silabo = con_silabo or 0
    finalizados = finalizados or 0
    total_asignaturas = total_asignaturas or 0
    con_plan = con_plan or 0
    total_carreras = total_carreras or 0
    kpi = {
        "total": total,
        "con_silabo": con_silabo,
        "sin_silabo": total - con_silabo,
        "finalizados": finalizados,
        "porcentaje": round((con_silabo / total) * 100) if total else 0,
        "total_asignaturas": total_asignaturas,
        "con_plan": con_plan,
        "sin_plan": total_asignaturas - con_plan,
        "porcentaje_plan": round((con_plan / total_asignaturas) * 100) if total_asignaturas else 0,
        "total_carreras": total_carreras,
    }

    # --- Gráfica ---
    if filtros["unidad_id"]:
        agrupar_id, agrupar_nombre = "c.carrera_id", "CASE WHEN c.malla IS NOT NULL THEN c.nombre + ' (' + c.malla + ')' ELSE c.nombre END"
    else:
        agrupar_id, agrupar_nombre = "ua.unidad_academica_id", "ua.nombre"

    cur.execute(f"""
        SELECT {agrupar_nombre},
               COUNT(oa.oferta_id) AS total,
               SUM(CASE WHEN s.existe_silabo = 1 THEN 1 ELSE 0 END) AS con_silabo,
               COUNT(DISTINCT m.materia_id) AS total_asignaturas,
               COUNT(DISTINCT CASE WHEN pa.existe_plan_analitico = 1 THEN m.materia_id END) AS con_plan
        {base_from} {where_sql}
        GROUP BY {agrupar_id}, {agrupar_nombre}
        ORDER BY {agrupar_nombre}
    """, params)
    filas_grafico = cur.fetchall()
    datos_grafico = {
        "labels": [r[0] for r in filas_grafico],
        "silabos": [round((r[2] / r[1]) * 100) if r[1] else 0 for r in filas_grafico],
        "asignaturas": [round((r[4] / r[3]) * 100) if r[3] else 0 for r in filas_grafico],
    }
    detalle_grupo = [
        {
            "nombre": r[0],
            "total": r[1],
            "porcentaje": round((r[2] / r[1]) * 100) if r[1] else 0,
            "total_asignaturas": r[3],
            "porcentaje_plan": round((r[4] / r[3]) * 100) if r[3] else 0,
        }
        for r in filas_grafico
    ]

    # --- Listas de estado ---
    listas = {
        "pendientes": _listar_materias(cur, base_from, where_sql, params, "(s.existe_silabo = 0 OR s.existe_silabo IS NULL)"),
        "por_finalizar": _listar_materias(cur, base_from, where_sql, params, "s.existe_silabo = 1 AND (s.finalizado = 0 OR s.finalizado IS NULL)"),
        "finalizados": _listar_materias(cur, base_from, where_sql, params, "s.finalizado = 1"),
        "sin_plan": _listar_materias(cur, base_from, where_sql, params, "(pa.existe_plan_analitico = 0 OR pa.existe_plan_analitico IS NULL)"),
    }

    # --- Tabla paginada ---
    if filtros["modo"] == "asignaturas":
        total_sql = f"SELECT COUNT(DISTINCT m.materia_id) {base_from} {where_sql}"
    elif filtros["modo"] == "silabos":
        total_sql = f"SELECT COUNT(oa.oferta_id) {base_from} {where_sql}"
    else:
        total_sql = f"SELECT COUNT(*) {base_from} {where_sql}"
    cur.execute(total_sql, params)
    total_filas = cur.fetchone()[0] or 0
    total_paginas = max(1, -(-total_filas // por_pagina))
    offset = (pagina - 1) * por_pagina

    if filtros["modo"] == "asignaturas":
        select_tabla = f"""
            SELECT ua.nombre,
                   CASE WHEN c.malla IS NOT NULL THEN c.nombre + ' (' + c.malla + ')' ELSE c.nombre END AS carrera_nombre,
                   m.cod_materia, m.nombre, NULL AS paralelo,
                   NULL AS profesor_curso, NULL AS profesor_silabo,
                   0 AS existe_silabo, 0 AS finalizado,
                   m.nivel, m.regimen_materia, ISNULL(pa.existe_plan_analitico, 0),
                   pa.ref_ultimo_plan_analitico, mme.periodo_academico_texto
            {base_from} {where_sql}
            GROUP BY ua.nombre, c.nombre, c.malla, m.materia_id, m.cod_materia, m.nombre,
                     m.nivel, m.regimen_materia, pa.existe_plan_analitico,
                     pa.ref_ultimo_plan_analitico, mme.periodo_academico_texto, mme.periodo_carga
            ORDER BY ua.nombre, carrera_nombre, m.nombre
            OFFSET ? ROWS FETCH NEXT ? ROWS ONLY
        """
    else:
        select_tabla = f"""
            SELECT ua.nombre, 
                   CASE WHEN c.malla IS NOT NULL THEN c.nombre + ' (' + c.malla + ')' ELSE c.nombre END AS carrera_nombre, 
                   m.cod_materia, m.nombre, oa.paralelo,
                   dc.nombre_completo, ds.nombre_completo,
                   ISNULL(s.existe_silabo, 0), ISNULL(s.finalizado, 0),
                   m.nivel, m.regimen_materia, ISNULL(pa.existe_plan_analitico, 0),
                   pa.ref_ultimo_plan_analitico, COALESCE(p.nombre, mme.periodo_academico_texto, mme.periodo_carga)
            {base_from} {where_sql}
            ORDER BY ua.nombre, carrera_nombre, m.nombre, oa.paralelo
            OFFSET ? ROWS FETCH NEXT ? ROWS ONLY
        """
    cur.execute(select_tabla, params + [offset, por_pagina])
    filas = [
        {
            "unidad": r[0], "carrera": r[1], "cod_materia": r[2], "materia": r[3], "paralelo": r[4],
            "profesor_curso": r[5], "profesor_silabo": r[6],
            "existe_silabo": bool(r[7]), "finalizado": bool(r[8]),
            "nivel": r[9], "regimen_materia": r[10],
            "existe_plan": bool(r[11]), "ref_plan": r[12],
            "periodo": r[13],
        }
        for r in cur.fetchall()
    ]

    # --- Listas para los selects de filtro ---
    cur.execute("""
        SELECT nombre FROM PeriodoAcademico
        UNION
        SELECT periodo FROM CargaArchivo WHERE periodo IS NOT NULL
        ORDER BY nombre
    """)
    periodos = [r[0] for r in cur.fetchall()]

    condicion_tipo, valores_tipo = _condicion_tipo_unidad(filtros.get("tipo_unidad", ""))
    where_tipo = f"WHERE {condicion_tipo}" if condicion_tipo and not filtros.get("unidad_id") else ""
    cur.execute(f"SELECT unidad_academica_id, nombre, tipo FROM UnidadAcademica ua {where_tipo} ORDER BY nombre", valores_tipo if where_tipo else [])
    unidades = [{"unidad_academica_id": r[0], "nombre": r[1], "tipo": r[2]} for r in cur.fetchall()]

    if filtros["unidad_id"]:
        cur.execute("SELECT carrera_id, CASE WHEN malla IS NOT NULL THEN nombre + ' (' + malla + ')' ELSE nombre END AS nombre_completo FROM Carrera WHERE unidad_academica_id = ? ORDER BY nombre", filtros["unidad_id"])
    elif where_tipo:
        cur.execute(f"""
            SELECT c.carrera_id, CASE WHEN c.malla IS NOT NULL THEN c.nombre + ' (' + c.malla + ')' ELSE c.nombre END AS nombre_completo
            FROM Carrera c
            JOIN UnidadAcademica ua ON ua.unidad_academica_id = c.unidad_academica_id
            {where_tipo}
            ORDER BY c.nombre
        """, valores_tipo)
    else:
        cur.execute("SELECT carrera_id, CASE WHEN malla IS NOT NULL THEN nombre + ' (' + malla + ')' ELSE nombre END AS nombre_completo FROM Carrera ORDER BY nombre")
    carreras = [{"carrera_id": r[0], "nombre": r[1]} for r in cur.fetchall()]

    conn.close()

    def url_pagina(n):
        args = {k: v for k, v in filtros.items() if v}
        args["pagina"] = n
        return "/dashboard?" + "&".join(f"{k}={v}" for k, v in args.items())

    conn2 = pyodbc.connect(CONN_STR)
    cur2 = conn2.cursor()
    ultima_carga = obtener_cargas_recientes(cur2, limite=1)
    conn2.close()

    return render_template(
        "dashboard.html", activo="inicio", iniciales=iniciales_usuario(),
        filtros=filtros, kpi=kpi, filas=filas, total_filas=total_filas,
        periodos=periodos, unidades=unidades, carreras=carreras,
        pagina=pagina, total_paginas=total_paginas, url_pagina=url_pagina,
        datos_grafico=json.dumps(datos_grafico), detalle_grupo=detalle_grupo,
        listas=listas, cargas=ultima_carga,
    )

def _construir_filtro_dashboard(filtros):
    condiciones, params = [], []
    if filtros["periodo"]:
        if filtros.get("modo") == "asignaturas":
            condiciones.append("""
                EXISTS (
                    SELECT 1
                    FROM MateriaMallaEstado mme_periodo
                    JOIN CargaArchivo ca_periodo ON ca_periodo.carga_id = mme_periodo.carga_id
                    WHERE mme_periodo.materia_id = m.materia_id
                      AND ca_periodo.periodo = ?
                )
            """)
            params.append(filtros["periodo"])
        elif filtros.get("modo") == "silabos":
            condiciones.append("p.nombre = ?"); params.append(filtros["periodo"])
        else:
            condiciones.append("""
                (p.nombre = ? OR EXISTS (
                    SELECT 1
                    FROM MateriaMallaEstado mme_periodo
                    JOIN CargaArchivo ca_periodo ON ca_periodo.carga_id = mme_periodo.carga_id
                    WHERE mme_periodo.materia_id = m.materia_id
                      AND ca_periodo.periodo = ?
                ))
            """)
            params.extend([filtros["periodo"], filtros["periodo"]])
    if filtros.get("tipo_unidad") and not filtros.get("unidad_id"):
        condicion, valores = _condicion_tipo_unidad(filtros["tipo_unidad"])
        if condicion:
            condiciones.append(condicion); params.extend(valores)
    if filtros["unidad_id"]:
        condiciones.append("ua.unidad_academica_id = ?"); params.append(filtros["unidad_id"])
    if filtros["carrera_id"]:
        condiciones.append("c.carrera_id = ?"); params.append(filtros["carrera_id"])
    if filtros["estado"] == "con_silabo":
        condiciones.append("s.existe_silabo = 1")
    elif filtros["estado"] == "sin_silabo":
        condiciones.append("(s.existe_silabo = 0 OR s.existe_silabo IS NULL)")
    elif filtros["estado"] == "finalizado":
        condiciones.append("s.finalizado = 1")
    elif filtros["estado"] == "por_finalizar":
        condiciones.append("s.existe_silabo = 1 AND (s.finalizado = 0 OR s.finalizado IS NULL)")
    elif filtros["estado"] == "sin_profesor":
        condiciones.append("oa.profesor_silabo_id IS NULL")
    elif filtros["estado"] == "con_plan":
        condiciones.append("pa.existe_plan_analitico = 1")
    elif filtros["estado"] == "sin_plan":
        condiciones.append("(pa.existe_plan_analitico = 0 OR pa.existe_plan_analitico IS NULL)")
    elif filtros["estado"] == "ofertada":
        condiciones.append("mme.ofertada_este_periodo = 1")
    elif filtros["estado"] == "no_ofertada":
        condiciones.append("(mme.ofertada_este_periodo = 0 OR mme.ofertada_este_periodo IS NULL)")
    if filtros["buscar"]:
        condiciones.append("(m.nombre LIKE ? OR m.cod_materia LIKE ? OR c.nombre LIKE ? OR dc.nombre_completo LIKE ? OR ISNULL(ds.nombre_completo,'') LIKE ? OR ISNULL(pa.ref_ultimo_plan_analitico,'') LIKE ?)")
        like = f"%{filtros['buscar']}%"
        params.extend([like, like, like, like, like, like])

    where_sql = ("WHERE " + " AND ".join(condiciones)) if condiciones else ""
    return where_sql, params

def _listar_materias(cur, base_from, where_sql, params, condicion_extra, limite=6):
    conector = "AND" if where_sql else "WHERE"
    sql = f"""
        SELECT DISTINCT TOP {limite} m.nombre, CASE WHEN c.malla IS NOT NULL THEN c.nombre + ' (' + c.malla + ')' ELSE c.nombre END
        {base_from} {where_sql} {conector} {condicion_extra}
        ORDER BY m.nombre
    """
    cur.execute(sql, params)
    return [f"{r[0]} — {r[1]}" for r in cur.fetchall()]

# ----------------------------------------------------------------------
# RUTAS: FLUJO DE SÍLABOS
# ----------------------------------------------------------------------
def _filtros_reporte():
    return {
        "periodo": request.args.get("periodo", ""),
        "tipo_unidad": request.args.get("tipo_unidad", ""),
        "unidad_id": request.args.get("unidad_id", ""),
        "carrera_id": request.args.get("carrera_id", ""),
        "estado": request.args.get("estado", ""),
        "buscar": request.args.get("buscar", ""),
    }

def _condicion_tipo_unidad(tipo_unidad):
    sedes = ["%BAHÍA%", "%BAHIA%", "%CHONE%", "%EL CARMEN%", "%PEDERNALES%", "%TOSAGUA%", "%SUCRE%", "%JAMA%", "%JIPIJAPA%", "%PICHINCHA%"]
    if tipo_unidad == "FACULTAD":
        return "ua.tipo = ? AND " + " AND ".join(["ua.nombre NOT LIKE ?" for _ in sedes]), ["FACULTAD"] + sedes
    if tipo_unidad == "SEDE_EXTENSION":
        return "(ua.tipo IN (?, ?) OR " + " OR ".join(["ua.nombre LIKE ?" for _ in sedes]) + ")", ["SEDE", "EXTENSION"] + sedes
    if tipo_unidad == "CAMPUS":
        return "ua.tipo = ?", ["CAMPUS"]
    return None, []

def _opciones_filtros(cur, filtros):
    cur.execute("""
        SELECT nombre FROM PeriodoAcademico
        UNION
        SELECT periodo FROM CargaArchivo WHERE periodo IS NOT NULL
        ORDER BY nombre
    """)
    periodos = [r[0] for r in cur.fetchall()]

    condicion_tipo, valores_tipo = _condicion_tipo_unidad(filtros.get("tipo_unidad", ""))
    where_tipo = f"WHERE {condicion_tipo}" if condicion_tipo and not filtros.get("unidad_id") else ""
    cur.execute(f"SELECT unidad_academica_id, nombre, tipo FROM UnidadAcademica ua {where_tipo} ORDER BY nombre", valores_tipo if where_tipo else [])
    unidades = [{"unidad_academica_id": r[0], "nombre": r[1], "tipo": r[2]} for r in cur.fetchall()]

    if filtros["unidad_id"]:
        cur.execute("""
            SELECT carrera_id,
                   CASE WHEN malla IS NOT NULL THEN nombre + ' (' + malla + ')' ELSE nombre END
            FROM Carrera
            WHERE unidad_academica_id = ?
            ORDER BY nombre
        """, filtros["unidad_id"])
    elif where_tipo:
        cur.execute(f"""
            SELECT c.carrera_id,
                   CASE WHEN c.malla IS NOT NULL THEN c.nombre + ' (' + c.malla + ')' ELSE c.nombre END
            FROM Carrera c
            JOIN UnidadAcademica ua ON ua.unidad_academica_id = c.unidad_academica_id
            {where_tipo}
            ORDER BY c.nombre
        """, valores_tipo)
    else:
        cur.execute("""
            SELECT carrera_id,
                   CASE WHEN malla IS NOT NULL THEN nombre + ' (' + malla + ')' ELSE nombre END
            FROM Carrera
            ORDER BY nombre
        """)
    carreras = [{"carrera_id": r[0], "nombre": r[1]} for r in cur.fetchall()]
    return periodos, unidades, carreras

def _csv_response(nombre_archivo, encabezados, filas):
    salida = io.StringIO()
    escritor = csv.writer(salida)
    escritor.writerow(encabezados)
    escritor.writerows(filas)
    contenido = "\ufeff" + salida.getvalue()
    return Response(
        contenido,
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename={nombre_archivo}"},
    )

def porcentaje(valor, total):
    return round((valor / total) * 100, 2) if total else 0

def fusionar_resumen_unidades(filas):
    acumulado = {}
    for fila in filas:
        unidad = normalizar_nombre_unidad(fila["unidad"])
        clave = clave_texto(unidad)
        item = acumulado.setdefault(clave, {"unidad": unidad, "ofertadas": 0, "creados": 0, "finalizados": 0})
        item["ofertadas"] += fila["ofertadas"]
        item["creados"] += fila["creados"]
        item["finalizados"] += fila["finalizados"]
    resultado = []
    for item in acumulado.values():
        item["porcentaje_creados"] = porcentaje(item["creados"], item["ofertadas"])
        item["porcentaje_finalizados"] = porcentaje(item["finalizados"], item["ofertadas"])
        resultado.append(item)
    return sorted(resultado, key=lambda x: x["unidad"])

def fusionar_resumen_carreras(filas):
    acumulado = {}
    for fila in filas:
        unidad = normalizar_nombre_unidad(fila["unidad"])
        carrera = normalizar_nombre_carrera(fila["carrera"])
        clave = (clave_texto(unidad), clave_texto(carrera))
        item = acumulado.setdefault(clave, {"unidad": unidad, "carrera": carrera, "ofertadas": 0, "creados": 0, "finalizados": 0})
        item["ofertadas"] += fila["ofertadas"]
        item["creados"] += fila["creados"]
        item["finalizados"] += fila["finalizados"]
    resultado = []
    for item in acumulado.values():
        item["porcentaje_creados"] = porcentaje(item["creados"], item["ofertadas"])
        item["porcentaje_finalizados"] = porcentaje(item["finalizados"], item["ofertadas"])
        resultado.append(item)
    return sorted(resultado, key=lambda x: (x["unidad"], x["carrera"]))

def _excel_response(nombre_archivo, titulo, encabezados, filas):
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"
    ws.append([titulo])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(encabezados))
    ws["A1"].font = Font(bold=True, size=14, color="1F2544")
    ws.append(encabezados)

    header_fill = PatternFill("solid", fgColor="1C2560")
    for cell in ws[2]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill

    for fila in filas:
        ws.append(fila)

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = ws.dimensions
    for col_idx, _ in enumerate(encabezados, start=1):
        letra = get_column_letter(col_idx)
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in ws[letra])
        ws.column_dimensions[letra].width = min(max(max_len + 2, 12), 45)

    salida = io.BytesIO()
    wb.save(salida)
    salida.seek(0)
    return Response(
        salida.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={nombre_archivo}"},
    )

def _excel_silabos_response(nombre_archivo, resumen_unidades, resumen_carreras, encabezados_detalle, filas_detalle, tabla_periodos_headers=None, tabla_periodos_rows=None):
    wb = Workbook()
    wb.remove(wb.active)
    header_fill = PatternFill("solid", fgColor="1C2560")

    hojas = [
        ("Consolidado unidades", ["Facultad/Extension/Sede/Campus", "Asignaturas ofertadas", "Silabos creados", "% creados", "Silabos finalizados", "% finalizados"], resumen_unidades),
        ("Consolidado carreras", ["Unidad", "Carrera / malla", "Asignaturas ofertadas", "Silabos creados", "% creados", "Silabos finalizados", "% finalizados"], resumen_carreras),
        ("Detalle", encabezados_detalle, filas_detalle),
    ]
    if tabla_periodos_headers and tabla_periodos_rows:
        hojas.insert(0, ("Porcentaje por carrera", tabla_periodos_headers, tabla_periodos_rows))
    for titulo, encabezados, filas in hojas:
        ws = wb.create_sheet(titulo[:31])
        ws.append([titulo])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(encabezados))
        ws["A1"].font = Font(bold=True, size=14, color="1F2544")
        ws.append(encabezados)
        for cell in ws[2]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
        for fila in filas:
            ws.append(fila)
        ws.freeze_panes = "A3"
        ws.auto_filter.ref = ws.dimensions
        for col_idx, _ in enumerate(encabezados, start=1):
            letra = get_column_letter(col_idx)
            max_len = max((len(str(c.value)) if c.value is not None else 0) for c in ws[letra])
            ws.column_dimensions[letra].width = min(max(max_len + 2, 12), 50)

    salida = io.BytesIO()
    wb.save(salida)
    salida.seek(0)
    return Response(
        salida.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={nombre_archivo}"},
    )

def _docx_p(texto, bold=False, size=22):
    texto = escape(str(texto or ""))
    bold_xml = "<w:b/>" if bold else ""
    return (
        "<w:p><w:r><w:rPr>"
        f"{bold_xml}<w:sz w:val=\"{size}\"/>"
        f"</w:rPr><w:t>{texto}</w:t></w:r></w:p>"
    )

def _docx_tbl(headers, rows):
    def cell(texto, bold=False):
        texto = escape(str(texto if texto is not None else ""))
        bold_xml = "<w:b/>" if bold else ""
        return (
            "<w:tc><w:tcPr><w:tcW w:w=\"2400\" w:type=\"dxa\"/></w:tcPr>"
            f"<w:p><w:r><w:rPr>{bold_xml}<w:sz w:val=\"18\"/></w:rPr><w:t>{texto}</w:t></w:r></w:p>"
            "</w:tc>"
        )

    header_row = "<w:tr>" + "".join(cell(h, True) for h in headers) + "</w:tr>"
    body_rows = "".join("<w:tr>" + "".join(cell(v) for v in row) + "</w:tr>" for row in rows)
    return (
        "<w:tbl><w:tblPr><w:tblW w:w=\"0\" w:type=\"auto\"/>"
        "<w:tblBorders><w:top w:val=\"single\" w:sz=\"4\"/><w:left w:val=\"single\" w:sz=\"4\"/>"
        "<w:bottom w:val=\"single\" w:sz=\"4\"/><w:right w:val=\"single\" w:sz=\"4\"/>"
        "<w:insideH w:val=\"single\" w:sz=\"4\"/><w:insideV w:val=\"single\" w:sz=\"4\"/>"
        "</w:tblBorders></w:tblPr>"
        f"{header_row}{body_rows}</w:tbl>"
    )

def _docx_response(nombre_archivo, titulo, secciones, filtros):
    periodo = filtros.get("periodo") or "Todos los periodos"
    body = [
        _docx_p(titulo, bold=True, size=30),
        _docx_p(f"Periodo: {periodo}", size=20),
        _docx_p("Fuente: Sistema de Gestion Academica Uleam", size=18),
        _docx_p("El calculo considera unicamente asignaturas ofertadas segun el periodo academico registrado en Asignaturas.", size=18),
    ]
    for seccion in secciones:
        body.append(_docx_p(seccion["titulo"], bold=True, size=24))
        body.append(_docx_tbl(seccion["headers"], seccion["rows"]))

    document_xml = (
        "<?xml version=\"1.0\" encoding=\"UTF-8\" standalone=\"yes\"?>"
        "<w:document xmlns:w=\"http://schemas.openxmlformats.org/wordprocessingml/2006/main\">"
        "<w:body>"
        + "".join(body)
        + "<w:sectPr><w:pgSz w:w=\"16838\" w:h=\"11906\" w:orient=\"landscape\"/>"
        "<w:pgMar w:top=\"720\" w:right=\"720\" w:bottom=\"720\" w:left=\"720\" w:header=\"360\" w:footer=\"360\" w:gutter=\"0\"/>"
        "</w:sectPr></w:body></w:document>"
    )
    content_types = (
        "<?xml version=\"1.0\" encoding=\"UTF-8\" standalone=\"yes\"?>"
        "<Types xmlns=\"http://schemas.openxmlformats.org/package/2006/content-types\">"
        "<Default Extension=\"rels\" ContentType=\"application/vnd.openxmlformats-package.relationships+xml\"/>"
        "<Default Extension=\"xml\" ContentType=\"application/xml\"/>"
        "<Override PartName=\"/word/document.xml\" ContentType=\"application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml\"/>"
        "</Types>"
    )
    rels = (
        "<?xml version=\"1.0\" encoding=\"UTF-8\" standalone=\"yes\"?>"
        "<Relationships xmlns=\"http://schemas.openxmlformats.org/package/2006/relationships\">"
        "<Relationship Id=\"rId1\" Type=\"http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument\" Target=\"word/document.xml\"/>"
        "</Relationships>"
    )
    salida = io.BytesIO()
    with zipfile.ZipFile(salida, "w", zipfile.ZIP_DEFLATED) as docx:
        docx.writestr("[Content_Types].xml", content_types)
        docx.writestr("_rels/.rels", rels)
        docx.writestr("word/document.xml", document_xml)
    salida.seek(0)
    return Response(
        salida.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": f"attachment; filename={nombre_archivo}"},
    )

def _construir_filtro_reporte_silabos(filtros, incluir_estado=True):
    condiciones, params = [], []
    if filtros["periodo"]:
        condiciones.append("p.nombre = ?"); params.append(filtros["periodo"])
    if filtros.get("tipo_unidad") and not filtros.get("unidad_id"):
        condicion, valores = _condicion_tipo_unidad(filtros["tipo_unidad"])
        if condicion:
            condiciones.append(condicion); params.extend(valores)
    if filtros["unidad_id"]:
        condiciones.append("ua.unidad_academica_id = ?"); params.append(filtros["unidad_id"])
    if filtros["carrera_id"]:
        condiciones.append("c.carrera_id = ?"); params.append(filtros["carrera_id"])
    if incluir_estado and filtros["estado"] == "con_silabo":
        condiciones.append("s.existe_silabo = 1")
    elif incluir_estado and filtros["estado"] == "sin_silabo":
        condiciones.append("(s.existe_silabo = 0 OR s.existe_silabo IS NULL)")
    elif incluir_estado and filtros["estado"] == "finalizado":
        condiciones.append("s.finalizado = 1")
    elif incluir_estado and filtros["estado"] == "por_finalizar":
        condiciones.append("s.existe_silabo = 1 AND (s.finalizado = 0 OR s.finalizado IS NULL)")
    elif incluir_estado and filtros["estado"] == "sin_profesor":
        condiciones.append("oa.profesor_silabo_id IS NULL")
    if filtros["buscar"]:
        condiciones.append("(m.nombre LIKE ? OR m.cod_materia LIKE ? OR c.nombre LIKE ? OR dc.nombre_completo LIKE ? OR ISNULL(ds.nombre_completo,'') LIKE ?)")
        like = f"%{filtros['buscar']}%"
        params.extend([like, like, like, like, like])
    return ("WHERE " + " AND ".join(condiciones)) if condiciones else "", params

def _agregar_condicion(where_sql, condicion):
    if where_sql:
        return f"{where_sql} AND {condicion}"
    return f"WHERE {condicion}"

def _construir_filtro_reporte_asignaturas(filtros):
    condiciones, params = [], []
    if filtros["periodo"]:
        condiciones.append("mme.periodo_carga = ?"); params.append(filtros["periodo"])
    if filtros.get("tipo_unidad") and not filtros.get("unidad_id"):
        condicion, valores = _condicion_tipo_unidad(filtros["tipo_unidad"])
        if condicion:
            condiciones.append(condicion); params.extend(valores)
    if filtros["unidad_id"]:
        condiciones.append("ua.unidad_academica_id = ?"); params.append(filtros["unidad_id"])
    if filtros["carrera_id"]:
        condiciones.append("c.carrera_id = ?"); params.append(filtros["carrera_id"])
    if filtros["estado"] == "con_plan":
        condiciones.append("pa.existe_plan_analitico = 1")
    elif filtros["estado"] == "sin_plan":
        condiciones.append("(pa.existe_plan_analitico = 0 OR pa.existe_plan_analitico IS NULL)")
    elif filtros["estado"] == "ofertada":
        condiciones.append("mme.ofertada_este_periodo = 1")
    elif filtros["estado"] == "no_ofertada":
        condiciones.append("(mme.ofertada_este_periodo = 0 OR mme.ofertada_este_periodo IS NULL)")
    if filtros["buscar"]:
        condiciones.append("(m.nombre LIKE ? OR m.cod_materia LIKE ? OR c.nombre LIKE ? OR ISNULL(pa.ref_ultimo_plan_analitico,'') LIKE ?)")
        like = f"%{filtros['buscar']}%"
        params.extend([like, like, like, like])
    return ("WHERE " + " AND ".join(condiciones)) if condiciones else "", params

@app.route("/", methods=["GET"])
@login_requerido
def index():
    conn = pyodbc.connect(CONN_STR)
    cur = conn.cursor()
    cargas = obtener_cargas_recientes(cur, tipo_reporte="SILABO")
    conn.close()
    return render_template("index.html", activo="silabos", iniciales=iniciales_usuario(), resultado=None, cargas=cargas)

@app.route("/silabos/reporte", methods=["GET"])
@login_requerido
def reporte_silabos():
    filtros = _filtros_reporte()
    descargar = request.args.get("descargar")
    conn = pyodbc.connect(CONN_STR)
    cur = conn.cursor()
    where_sql, params = _construir_filtro_reporte_silabos(filtros)
    where_resumen_sql, params_resumen = _construir_filtro_reporte_silabos(filtros, incluir_estado=False)
    base_from = """
        FROM OfertaAcademica oa
        JOIN Materia m ON m.materia_id = oa.materia_id
        JOIN Carrera c ON c.carrera_id = m.carrera_id
        JOIN UnidadAcademica ua ON ua.unidad_academica_id = c.unidad_academica_id
        JOIN PeriodoAcademico p ON p.periodo_id = oa.periodo_id
        LEFT JOIN Docente dc ON dc.docente_id = oa.profesor_curso_id
        LEFT JOIN Docente ds ON ds.docente_id = oa.profesor_silabo_id
        LEFT JOIN Silabo s ON s.oferta_id = oa.oferta_id
        OUTER APPLY (
            SELECT TOP 1 estado.ofertada_este_periodo
            FROM MateriaMallaEstado estado
            JOIN CargaArchivo ca ON ca.carga_id = estado.carga_id
            WHERE estado.materia_id = m.materia_id
              AND (ca.periodo = p.nombre OR estado.periodo_academico_texto = p.nombre)
            ORDER BY estado.fecha_registro DESC, estado.estado_id DESC
        ) mme
    """
    where_resumen_sql = _agregar_condicion(where_resumen_sql, "ISNULL(mme.ofertada_este_periodo, 0) = 1")
    where_detalle_sql = _agregar_condicion(where_sql, "ISNULL(mme.ofertada_este_periodo, 0) = 1")

    cur.execute(f"""
        SELECT COUNT(*),
               SUM(CASE WHEN s.existe_silabo = 1 THEN 1 ELSE 0 END),
               SUM(CASE WHEN s.finalizado = 1 THEN 1 ELSE 0 END),
               SUM(CASE WHEN s.existe_silabo = 1 AND (s.finalizado = 0 OR s.finalizado IS NULL) THEN 1 ELSE 0 END)
        {base_from} {where_detalle_sql}
    """, params)
    total, con_silabo, finalizados, por_finalizar = cur.fetchone()
    total = total or 0
    con_silabo = con_silabo or 0
    finalizados = finalizados or 0
    por_finalizar = por_finalizar or 0
    kpi = {
        "total": total,
        "con_silabo": con_silabo,
        "sin_silabo": total - con_silabo,
        "finalizados": finalizados,
        "por_finalizar": por_finalizar,
        "porcentaje": round((con_silabo / total) * 100) if total else 0,
        "porcentaje_finalizado": round((finalizados / total) * 100) if total else 0,
    }

    cur.execute(f"""
        SELECT ua.nombre,
               COUNT(*),
               SUM(CASE WHEN s.existe_silabo = 1 THEN 1 ELSE 0 END),
               SUM(CASE WHEN s.finalizado = 1 THEN 1 ELSE 0 END)
        {base_from} {where_resumen_sql}
        GROUP BY ua.nombre
        ORDER BY ua.nombre
    """, params_resumen)
    resumen_unidades = []
    for r in cur.fetchall():
        ofertadas = r[1] or 0
        creados = r[2] or 0
        finalizados_resumen = r[3] or 0
        resumen_unidades.append({
            "unidad": r[0],
            "ofertadas": ofertadas,
            "creados": creados,
            "finalizados": finalizados_resumen,
            "porcentaje_creados": porcentaje(creados, ofertadas),
            "porcentaje_finalizados": porcentaje(finalizados_resumen, ofertadas),
        })
    resumen_unidades = fusionar_resumen_unidades(resumen_unidades)

    cur.execute(f"""
        SELECT ua.nombre,
               CASE WHEN c.malla IS NOT NULL THEN c.nombre + ' (' + c.malla + ')' ELSE c.nombre END,
               COUNT(*),
               SUM(CASE WHEN s.existe_silabo = 1 THEN 1 ELSE 0 END),
               SUM(CASE WHEN s.finalizado = 1 THEN 1 ELSE 0 END)
        {base_from} {where_resumen_sql}
        GROUP BY ua.nombre, c.nombre, c.malla
        ORDER BY ua.nombre, c.nombre
    """, params_resumen)
    resumen_carreras = []
    for r in cur.fetchall():
        ofertadas = r[2] or 0
        creados = r[3] or 0
        finalizados_resumen = r[4] or 0
        resumen_carreras.append({
            "unidad": r[0],
            "carrera": normalizar_nombre_carrera(r[1]),
            "ofertadas": ofertadas,
            "creados": creados,
            "finalizados": finalizados_resumen,
            "porcentaje_creados": porcentaje(creados, ofertadas),
            "porcentaje_finalizados": porcentaje(finalizados_resumen, ofertadas),
        })
    resumen_carreras = fusionar_resumen_carreras(resumen_carreras)

    cur.execute(f"""
        SELECT p.nombre,
               ua.nombre,
               COUNT(*),
               SUM(CASE WHEN s.existe_silabo = 1 THEN 1 ELSE 0 END),
               SUM(CASE WHEN s.finalizado = 1 THEN 1 ELSE 0 END)
        {base_from} {where_resumen_sql}
        GROUP BY p.nombre, ua.nombre
        ORDER BY p.nombre, ua.nombre
    """, params_resumen)
    resumen_por_periodo = {}
    for r in cur.fetchall():
        ofertadas = r[2] or 0
        creados = r[3] or 0
        finalizados_resumen = r[4] or 0
        resumen_por_periodo.setdefault(r[0], []).append([
            r[1], ofertadas, creados, f"{porcentaje(creados, ofertadas):.2f}",
            finalizados_resumen, f"{porcentaje(finalizados_resumen, ofertadas):.2f}",
        ])

    cur.execute(f"""
        SELECT ua.nombre,
               CASE WHEN c.malla IS NOT NULL THEN c.nombre + ' (' + c.malla + ')' ELSE c.nombre END,
               p.nombre,
               COUNT(*),
               SUM(CASE WHEN s.finalizado = 1 THEN 1 ELSE 0 END)
        {base_from} {where_resumen_sql}
        GROUP BY ua.nombre, c.nombre, c.malla, p.nombre
        ORDER BY ua.nombre, c.nombre, p.nombre
    """, params_resumen)
    periodos_comparativo = []
    tabla_periodos_map = {}
    for r in cur.fetchall():
        periodo_nombre = r[2]
        if periodo_nombre not in periodos_comparativo:
            periodos_comparativo.append(periodo_nombre)
        unidad_normalizada = normalizar_nombre_unidad(r[0])
        carrera_normalizada = normalizar_nombre_carrera(r[1])
        clave = (clave_texto(unidad_normalizada), clave_texto(carrera_normalizada))
        ofertadas = r[3] or 0
        finalizados_periodo = r[4] or 0
        item = tabla_periodos_map.setdefault(clave, {"unidad": unidad_normalizada, "carrera": carrera_normalizada, "periodos": {}})
        acumulado = item["periodos"].setdefault(periodo_nombre, {"ofertadas": 0, "finalizados": 0})
        acumulado["ofertadas"] += ofertadas
        acumulado["finalizados"] += finalizados_periodo
    tabla_periodos = [
        {
            "unidad": item["unidad"],
            "carrera": item["carrera"],
            "periodos": [
                porcentaje(item["periodos"][p]["finalizados"], item["periodos"][p]["ofertadas"]) if p in item["periodos"] else None
                for p in periodos_comparativo
            ],
        }
        for item in sorted(tabla_periodos_map.values(), key=lambda x: (x["unidad"], x["carrera"]))
    ]
    tabla_periodos_headers = ["Unidad Academica", "Carrera"] + periodos_comparativo
    tabla_periodos_rows = [
        [r["unidad"], r["carrera"]] + ["" if v is None else f"{v:.2f}" for v in r["periodos"]]
        for r in tabla_periodos
    ]

    cur.execute(f"""
        SELECT ua.nombre,
               CASE WHEN c.malla IS NOT NULL THEN c.nombre + ' (' + c.malla + ')' ELSE c.nombre END,
               p.nombre, m.cod_materia, m.nombre, oa.paralelo,
               dc.nombre_completo, ds.nombre_completo,
               ISNULL(s.existe_silabo, 0), ISNULL(s.finalizado, 0)
        {base_from} {where_detalle_sql}
        ORDER BY ua.nombre, c.nombre, m.nombre, oa.paralelo
    """, params)
    filas_raw = cur.fetchall()
    filas = [
        {
            "unidad": r[0], "carrera": r[1], "periodo": r[2], "cod_materia": r[3],
            "materia": r[4], "paralelo": r[5], "profesor_curso": r[6],
            "profesor_silabo": r[7], "existe_silabo": bool(r[8]), "finalizado": bool(r[9]),
        }
        for r in filas_raw
    ]

    encabezados = ["Unidad", "Carrera", "Periodo", "Codigo", "Materia", "Paralelo", "Profesor curso", "Profesor silabo", "Existe silabo", "Finalizado"]
    filas_export = [[f["unidad"], f["carrera"], f["periodo"], f["cod_materia"], f["materia"], f["paralelo"], f["profesor_curso"], f["profesor_silabo"], "Si" if f["existe_silabo"] else "No", "Si" if f["finalizado"] else "No"] for f in filas]
    if descargar == "csv":
        conn.close()
        return _csv_response("reporte_silabos.csv", encabezados, filas_export)
    if descargar == "xlsx":
        resumen_unidades_export = [
            [r["unidad"], r["ofertadas"], r["creados"], r["porcentaje_creados"], r["finalizados"], r["porcentaje_finalizados"]]
            for r in resumen_unidades
        ]
        resumen_carreras_export = [
            [r["unidad"], r["carrera"], r["ofertadas"], r["creados"], r["porcentaje_creados"], r["finalizados"], r["porcentaje_finalizados"]]
            for r in resumen_carreras
        ]
        conn.close()
        return _excel_silabos_response("reporte_silabos.xlsx", resumen_unidades_export, resumen_carreras_export, encabezados, filas_export, tabla_periodos_headers, tabla_periodos_rows)
    if descargar == "docx":
        headers_unidades = ["Facultad/Extension/Sede/Campus", "Asignaturas ofertadas", "Silabos creados", "% creados", "Silabos finalizados", "% finalizados"]
        secciones = []
        if tabla_periodos_rows:
            secciones.append({
                "titulo": "Cuadro comparativo de porcentaje de silabos finalizados por carrera",
                "headers": tabla_periodos_headers,
                "rows": tabla_periodos_rows,
            })
        if filtros["periodo"]:
            secciones.append({
                "titulo": f"Cuadro comparativo de silabos por asignaturas, creados y finalizados {filtros['periodo']}",
                "headers": headers_unidades,
                "rows": [
                    [r["unidad"], r["ofertadas"], r["creados"], f"{r['porcentaje_creados']:.2f}", r["finalizados"], f"{r['porcentaje_finalizados']:.2f}"]
                    for r in resumen_unidades
                ],
            })
        else:
            for periodo_nombre, filas_periodo in resumen_por_periodo.items():
                secciones.append({
                    "titulo": f"Cuadro comparativo de silabos por asignaturas, creados y finalizados {periodo_nombre}",
                    "headers": headers_unidades,
                    "rows": filas_periodo,
                })
        secciones.append({
            "titulo": "Cuadro comparativo por carrera / malla",
            "headers": ["Unidad", "Carrera / malla", "Ofertadas", "Creados", "% creados", "Finalizados", "% finalizados"],
            "rows": [
                [r["unidad"], r["carrera"], r["ofertadas"], r["creados"], f"{r['porcentaje_creados']:.2f}", r["finalizados"], f"{r['porcentaje_finalizados']:.2f}"]
                for r in resumen_carreras
            ],
        })
        conn.close()
        return _docx_response("informe_silabos.docx", "Informe tecnico de seguimiento al silabo", secciones, filtros)

    periodos, unidades, carreras = _opciones_filtros(cur, filtros)
    conn.close()
    return render_template(
        "reporte_silabos.html", activo="silabos", iniciales=iniciales_usuario(),
        filtros=filtros, kpi=kpi, filas=filas, periodos=periodos, unidades=unidades,
        carreras=carreras, resumen_unidades=resumen_unidades, resumen_carreras=resumen_carreras,
        periodos_comparativo=periodos_comparativo, tabla_periodos=tabla_periodos,
    )

@app.route("/cargar", methods=["POST"])
@login_requerido
def cargar():
    archivo = request.files.get("archivo")
    usuario = session.get("nombre_usuario", "desconocido")

    if not archivo or archivo.filename == "":
        return _renderizar_con_resultado({"estado": "ERROR", "mensaje_error": "No se seleccionó ningún archivo."})

    if not archivo.filename.lower().endswith((".xlsx", ".xls")):
        return _renderizar_con_resultado({"estado": "ERROR", "mensaje_error": "El archivo debe ser .xlsx o .xls."})

    with tempfile.TemporaryDirectory() as tmp:
        ruta_temp = os.path.join(tmp, archivo.filename)
        archivo.save(ruta_temp)
        diagnostico = diagnosticar_archivo_carga(ruta_temp, "SILABOS")
        if not diagnostico["valido"]:
            registrar_error_validacion(archivo.filename, "SILABO", diagnostico["mensaje"])
            return _renderizar_con_resultado({
                "estado": "ERROR",
                "mensaje_error": diagnostico["mensaje"],
                "diagnostico": diagnostico,
            })

        conn = pyodbc.connect(CONN_STR)
        cur = conn.cursor()
        try:
            periodo, unidad, contadores = migrar_archivo(cur, ruta_temp, archivo.filename, usuario)
            registrar_carga(cur, archivo.filename, periodo, unidad, contadores, "EXITOSO")
            conn.commit()
            resultado = {
                "estado": "EXITOSO", "periodo": periodo, "unidad": unidad,
                "contadores": contadores, "diagnostico": diagnostico,
            }
        except Exception as e:
            conn.rollback()
            mensaje = str(e) if isinstance(e, ValueError) else "Ocurrió un error inesperado."
            try:
                registrar_carga(cur, archivo.filename, None, None, None, "ERROR", mensaje)
                conn.commit()
            except Exception: pass
            traceback.print_exc()
            resultado = {"estado": "ERROR", "mensaje_error": mensaje}
        finally:
            conn.close()

    return _renderizar_con_resultado(resultado)

def _renderizar_con_resultado(resultado):
    conn = pyodbc.connect(CONN_STR)
    cur = conn.cursor()
    cargas = obtener_cargas_recientes(cur, tipo_reporte="SILABO")
    conn.close()
    return render_template("index.html", activo="silabos", iniciales=iniciales_usuario(), resultado=resultado, cargas=cargas)


# ----------------------------------------------------------------------
# RUTAS: FLUJO DE PROGRAMA ANALÍTICO
# ----------------------------------------------------------------------
@app.route("/programa-analitico", methods=["GET"])
@login_requerido
def programa_analitico():
    conn = pyodbc.connect(CONN_STR)
    cur = conn.cursor()
    cargas = obtener_cargas_recientes(cur, tipo_reporte="PROGRAMA_ANALITICO")
    conn.close()
    return render_template(
        "programa_analitico.html", activo="programa_analitico", iniciales=iniciales_usuario(),
        resultado=None, cargas=cargas,
    )

@app.route("/asignaturas/reporte", methods=["GET"])
@login_requerido
def reporte_asignaturas():
    filtros = _filtros_reporte()
    descargar = request.args.get("descargar")
    conn = pyodbc.connect(CONN_STR)
    cur = conn.cursor()
    where_sql, params = _construir_filtro_reporte_asignaturas(filtros)
    params_base = [filtros["periodo"], filtros["periodo"]]
    base_from = """
        FROM Materia m
        JOIN Carrera c ON c.carrera_id = m.carrera_id
        JOIN UnidadAcademica ua ON ua.unidad_academica_id = c.unidad_academica_id
        LEFT JOIN PlanAnalitico pa ON pa.materia_id = m.materia_id
        OUTER APPLY (
            SELECT TOP 1 estado.periodo_academico_texto, estado.ofertada_este_periodo,
                         ca.periodo AS periodo_carga, ca.fecha_carga
            FROM MateriaMallaEstado estado
            JOIN CargaArchivo ca ON ca.carga_id = estado.carga_id
            WHERE estado.materia_id = m.materia_id
              AND (? = '' OR ca.periodo = ?)
            ORDER BY estado.fecha_registro DESC, estado.estado_id DESC
        ) mme
    """

    cur.execute(f"""
        SELECT COUNT(*),
               SUM(CASE WHEN pa.existe_plan_analitico = 1 THEN 1 ELSE 0 END),
               SUM(CASE WHEN mme.ofertada_este_periodo = 1 THEN 1 ELSE 0 END),
               COUNT(DISTINCT c.carrera_id)
        {base_from} {where_sql}
    """, params_base + params)
    total, con_plan, ofertadas, carreras_total = cur.fetchone()
    total = total or 0
    con_plan = con_plan or 0
    ofertadas = ofertadas or 0
    carreras_total = carreras_total or 0
    kpi = {
        "total": total,
        "con_plan": con_plan,
        "sin_plan": total - con_plan,
        "ofertadas": ofertadas,
        "carreras": carreras_total,
        "porcentaje_plan": round((con_plan / total) * 100) if total else 0,
    }

    cur.execute(f"""
        SELECT ua.nombre,
               CASE WHEN c.malla IS NOT NULL THEN c.nombre + ' (' + c.malla + ')' ELSE c.nombre END,
               m.nivel, m.cod_materia, m.nombre, m.regimen_materia,
               ISNULL(pa.existe_plan_analitico, 0), pa.ref_ultimo_plan_analitico,
               pa.observacion,
               mme.periodo_academico_texto,
               ISNULL(mme.ofertada_este_periodo, 0)
        {base_from} {where_sql}
        ORDER BY ua.nombre, c.nombre, m.nivel, m.nombre
    """, params_base + params)
    filas_raw = cur.fetchall()
    filas = [
        {
            "unidad": r[0], "carrera": r[1], "nivel": r[2], "cod_materia": r[3],
            "materia": r[4], "regimen_materia": r[5], "existe_plan": bool(r[6]),
            "ref_plan": r[7], "observacion": r[8], "periodo": r[9],
            "ofertada": bool(r[10]),
        }
        for r in filas_raw
    ]

    encabezados = ["Unidad", "Carrera", "Nivel", "Codigo", "Asignatura", "Regimen", "Existe plan", "Ref plan", "Observacion", "Periodo academico", "Ofertada"]
    filas_export = [[f["unidad"], f["carrera"], f["nivel"], f["cod_materia"], f["materia"], f["regimen_materia"], "Si" if f["existe_plan"] else "No", f["ref_plan"], f["observacion"], f["periodo"], "Si" if f["ofertada"] else "No"] for f in filas]
    if descargar == "csv":
        conn.close()
        return _csv_response("reporte_asignaturas.csv", encabezados, filas_export)
    if descargar == "xlsx":
        conn.close()
        return _excel_response("reporte_asignaturas.xlsx", "Reporte de Asignaturas", encabezados, filas_export)

    periodos, unidades, carreras = _opciones_filtros(cur, filtros)
    conn.close()
    return render_template(
        "reporte_asignaturas.html", activo="programa_analitico", iniciales=iniciales_usuario(),
        filtros=filtros, kpi=kpi, filas=filas, periodos=periodos, unidades=unidades,
        carreras=carreras,
    )

def _materias_de_carga(cur, carga_id):
    cur.execute("""
        SELECT c.ref_carrera, m.ref_materia, m.cod_materia, m.nombre, m.nivel,
               m.regimen_materia, c.nombre, c.malla, ua.nombre
        FROM MateriaMallaEstado mme
        JOIN Materia m ON m.materia_id = mme.materia_id
        JOIN Carrera c ON c.carrera_id = m.carrera_id
        JOIN UnidadAcademica ua ON ua.unidad_academica_id = c.unidad_academica_id
        WHERE mme.carga_id = ?
    """, carga_id)
    materias = {}
    for r in cur.fetchall():
        key = (r[0], r[1])
        materias[key] = {
            "ref_carrera": r[0], "ref_materia": r[1], "cod_materia": r[2],
            "materia": r[3], "nivel": r[4], "regimen": r[5],
            "carrera": f"{r[6]} ({r[7]})" if r[7] else r[6],
            "unidad": r[8],
        }
    return materias

@app.route("/asignaturas/cambios", methods=["GET"])
@login_requerido
def cambios_malla():
    conn = pyodbc.connect(CONN_STR)
    cur = conn.cursor()
    cur.execute("""
        SELECT carga_id, nombre_archivo, fecha_carga, periodo, unidad_academica
        FROM CargaArchivo
        WHERE tipo_reporte = 'PROGRAMA_ANALITICO' AND estado = 'EXITOSO'
        ORDER BY fecha_carga DESC
    """)
    cargas = [
        {"carga_id": r[0], "nombre_archivo": r[1], "fecha_carga": r[2], "periodo": r[3], "unidad": r[4]}
        for r in cur.fetchall()
    ]
    carga_id = request.args.get("carga_id", type=int) or (cargas[0]["carga_id"] if cargas else None)
    actual = next((c for c in cargas if c["carga_id"] == carga_id), None)
    anterior = None
    agregadas, retiradas, modificadas = [], [], []

    if actual:
        cur.execute("""
            SELECT TOP 1 carga_id, nombre_archivo, fecha_carga, periodo, unidad_academica
            FROM CargaArchivo
            WHERE tipo_reporte = 'PROGRAMA_ANALITICO'
              AND estado = 'EXITOSO'
              AND unidad_academica = ?
              AND fecha_carga < ?
            ORDER BY fecha_carga DESC
        """, actual["unidad"], actual["fecha_carga"])
        row = cur.fetchone()
        if row:
            anterior = {"carga_id": row[0], "nombre_archivo": row[1], "fecha_carga": row[2], "periodo": row[3], "unidad": row[4]}
            materias_actual = _materias_de_carga(cur, actual["carga_id"])
            materias_anterior = _materias_de_carga(cur, anterior["carga_id"])
            actuales = set(materias_actual.keys())
            anteriores = set(materias_anterior.keys())
            agregadas = [materias_actual[k] for k in sorted(actuales - anteriores)]
            retiradas = [materias_anterior[k] for k in sorted(anteriores - actuales)]
            for key in sorted(actuales & anteriores):
                antes = materias_anterior[key]
                despues = materias_actual[key]
                cambios = []
                for campo, etiqueta in [("cod_materia", "Código"), ("materia", "Asignatura"), ("nivel", "Nivel"), ("regimen", "Régimen"), ("carrera", "Carrera / malla")]:
                    if antes[campo] != despues[campo]:
                        cambios.append(f"{etiqueta}: {antes[campo] or '-'} -> {despues[campo] or '-'}")
                if cambios:
                    item = dict(despues)
                    item["cambios"] = cambios
                    modificadas.append(item)

    conn.close()
    return render_template(
        "cambios_malla.html", activo="cambios_malla", iniciales=iniciales_usuario(),
        cargas=cargas, actual=actual, anterior=anterior,
        agregadas=agregadas, retiradas=retiradas, modificadas=modificadas,
    )

@app.route("/historial", methods=["GET"])
@login_requerido
def historial():
    tipo = request.args.get("tipo", "")
    estado = request.args.get("estado", "")
    condiciones, params = [], []
    if tipo:
        condiciones.append("tipo_reporte = ?"); params.append(tipo)
    if estado:
        condiciones.append("estado = ?"); params.append(estado)
    where_sql = ("WHERE " + " AND ".join(condiciones)) if condiciones else ""
    conn = pyodbc.connect(CONN_STR)
    cur = conn.cursor()
    cur.execute(f"""
        SELECT TOP 100 nombre_archivo, fecha_carga, periodo, unidad_academica,
               registros_creados, registros_actualizados, registros_sin_cambio,
               estado, mensaje_error, tipo_reporte
        FROM CargaArchivo
        {where_sql}
        ORDER BY fecha_carga DESC
    """, params)
    columnas = [c[0] for c in cur.description]
    cargas = [dict(zip(columnas, fila)) for fila in cur.fetchall()]
    conn.close()
    return render_template(
        "historial.html", activo="historial", iniciales=iniciales_usuario(),
        cargas=cargas, tipo=tipo, estado=estado,
    )

@app.route("/cargar-analitico", methods=["POST"])
@login_requerido
def cargar_analitico():
    archivo = request.files.get("archivo")

    if not archivo or archivo.filename == "":
        return _renderizar_analitico_con_resultado({"estado": "ERROR", "mensaje_error": "No se seleccionó ningún archivo."})

    if not archivo.filename.lower().endswith((".xlsx", ".xls")):
        return _renderizar_analitico_con_resultado({"estado": "ERROR", "mensaje_error": "El archivo debe ser .xlsx o .xls."})

    with tempfile.TemporaryDirectory() as tmp:
        ruta_temp = os.path.join(tmp, archivo.filename)
        archivo.save(ruta_temp)
        diagnostico = diagnosticar_archivo_carga(ruta_temp, "ASIGNATURAS")
        if not diagnostico["valido"]:
            registrar_error_validacion(archivo.filename, "PROGRAMA_ANALITICO", diagnostico["mensaje"])
            return _renderizar_analitico_con_resultado({
                "estado": "ERROR",
                "mensaje_error": diagnostico["mensaje"],
                "diagnostico": diagnostico,
            })

        conn = pyodbc.connect(CONN_STR)
        cur = conn.cursor()
        carga_id = crear_carga_pendiente(cur, archivo.filename, "PROGRAMA_ANALITICO")
        conn.commit()
        
        try:
            periodo, unidad, contadores = migrar_archivo_analitico(cur, ruta_temp, archivo.filename, carga_id)
            finalizar_carga(cur, carga_id, periodo, unidad, contadores, "EXITOSO")
            conn.commit()
            resultado = {
                "estado": "EXITOSO", "periodo": periodo, "unidad": unidad,
                "contadores": contadores, "diagnostico": diagnostico,
            }
        except Exception as e:
            conn.rollback()
            mensaje = str(e) if isinstance(e, ValueError) else "Ocurrió un error inesperado procesando el archivo."
            try:
                finalizar_carga(cur, carga_id, None, None, None, "ERROR", mensaje)
                conn.commit()
            except Exception:
                conn.rollback()
            traceback.print_exc()
            resultado = {"estado": "ERROR", "mensaje_error": mensaje}
        finally:
            conn.close()

    return _renderizar_analitico_con_resultado(resultado)

def _renderizar_analitico_con_resultado(resultado):
    conn = pyodbc.connect(CONN_STR)
    cur = conn.cursor()
    cargas = obtener_cargas_recientes(cur, tipo_reporte="PROGRAMA_ANALITICO")
    conn.close()
    return render_template(
        "programa_analitico.html", activo="programa_analitico", iniciales=iniciales_usuario(),
        resultado=resultado, cargas=cargas,
    )

if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5000)
